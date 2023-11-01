# Imports
from os import path
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
import tkinter as tk
from tkinter import Label, Entry, Button, PhotoImage
from tkinter import font
from tkinter import messagebox
from tkinter import filedialog
from tkinter import YES, NO, END, LEFT, RIGHT
from PIL import Image, ImageTk
import datetime
from itertools import count, cycle

##################################################################################################################

# Constants
main_window_background_color = "#111"
title_background_color = "#111"
title_foreground_color = "#aaa"
label_background_color = "#111"
label_foreground_color = "#777"
textbox_background_color = "#bec1c2"
is_paper_checked = False
loading_time = 5000  # number in milliseconds
error = "#d92323"
correct_text = "lightgreen"
correct_check = "green"
normal_check_text = "#fff"
main_window_size = "465x780"
loading_screen_image_path = "./loadingImage.gif"
loading_screen_image = Image.open(loading_screen_image_path)
icon_img_path = './order_icon.ico'
paper_checkbox_background_color = "#333"
paper_checkbox_foreground_color = "red" # alternative color hex value: "#2f9e24"
code_checkbox_background_color = "#333"
code_checkbox_foreground_color = "#3361a6" # alternative color hex value: "#2f9e24"
saving_file_path = "./" # determined during runtime
header_text_background_color = "262626"
header_text_foreground_color = "ffffff"
review_and_code_excel_cell_color = "a9f73b"
review_only_excel_cell_color = "c93892"
code_only_excel_cell_color = "8db4e2"
excel_cell_color = "eaeaea"
button_font_size = 12
check_button_background_color = "#3a4245"
check_button_foreground_color = "#fff"
add_button_background_color = "#3a4245"
add_button_foreground_color = "#fff"



##################################################################################################################

# Functions
def create_label(content):
    label_font = font.Font(size=content["label_size"])
    label = Label(main_window, text=content["label_text"], font=label_font)
    label['background'] = content["label_background_color"]
    label['foreground'] = content["label_foreground_color"]
    return label


def create_textbox(content):
    textbox = Entry(main_window, width=20, font=("Helvetica", 20))
    textbox['background'] = content["textbox_background_color"]
    return textbox

def create_saving_location():
    global saving_file_path
    check_file_exist = messagebox.askyesno("Prompt", "Does the check File exist?")
    if check_file_exist:
        saving_file_path = filedialog.askopenfilename()
    else:
        choose_folder = filedialog.askdirectory()
        saving_file_path = create_file_path(choose_folder)
    return saving_file_path

def create_file_path(folder):
    file_path = folder + "/Research_Papers_for_" + paperReleaseDateTextbox.get() + ".xlsx"
    return file_path


def check_if_excel_file_exists(file_path):
    print(file_path)
    return path.exists(file_path)


def create_new_excel_file(file_path):
    # Create a new workbook
    workbook = Workbook()
    # Get the active sheet
    sheet = workbook.active
    # Specify the column names
    column_names = ["Research Paper Name", "Research Paper Link", "Model Accuracy", "Model Dataset",
                    "Model Algorithm", "Paper Release Date", "Notes", "Review Paper", "Code"]
    # Write the column names to the first row
    sheet.append(column_names)
    # Set the height of row 2 to 30
    row_index = 1
    row_dimensions = sheet.row_dimensions[row_index]
    row_dimensions.height = 30
    # Change Column Width
    for column_index in range(1, 10):
        column_letter = get_column_letter(column_index)
        column_dimensions = sheet.column_dimensions[column_letter]
        column_dimensions.width = 25
    # Set the content alignment for the entire column
    for cell in sheet.iter_cols(min_col=1, max_col=9):
        for col in cell:
            col.alignment = Alignment(horizontal="center", vertical="center")
    # Specify the range of columns and row index
    start_column = 1
    end_column = 9
    row_index = 1
    # Create a border style
    border = Border(
        top=Side(border_style="thick", color="000000"),
        bottom=Side(border_style="thick", color="000000"),
        left=Side(border_style="thick", color="000000"),
        right=Side(border_style="thick", color="000000")
    )
    # Set the background color for cells in the specified range
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(row=row_index, column=column)
        cell.fill = PatternFill(start_color=header_text_background_color, end_color=header_text_background_color,
                                fill_type="solid")
        cell.border = border
        cell.font = Font(bold=True, color=header_text_foreground_color)

    # Save the workbook to the specified file path
    workbook.save(file_path)


def check_empty_values():
    check = False

    for value in values:
        if change_background_color(value) is True:
            check = True
    return check


def change_background_color(value):
    value_length = len(value.get().strip())
    if value_length == 0:
        value.config(bg=error)
    else:
        value.config(bg=correct_text)
    return value_length == 0


def paperNameTextbox_background_color(*args):
    change_background_color(paperNameTextbox)


def paperLinkTextbox_background_color(*args):
    change_background_color(paperLinkTextbox)


def modelAccuracyTextbox_background_color(*args):
    change_background_color(modelAccuracyTextbox)


def modelDatasetTextbox_background_color(*args):
    change_background_color(modelDatasetTextbox)


def modelAlgorithmTextbox_background_color(*args):
    change_background_color(modelAlgorithmTextbox)


def paperReleaseDateTextbox_background_color(*args):
    change_background_color(paperReleaseDateTextbox)
    
def notesTextbox_background_color(*args):
    change_background_color(notesTextbox)


def check_if_name_and_link_repeated(file_path, name, link):
    # define the check variable which will be used as return value
    check = False
    
    # Define variable to load the dataframe
    dataframe = load_workbook(file_path)

    # Define variable to read sheet
    df = dataframe.active

    # Specify the range of columns and row index
    start_col = 1
    end_col = 2
    start_row = 2
    end_row = len(df['A'])

    # Set the new values for cells in the specified range
    for row in range(start_row, end_row + 1):
        for column in range(start_col, end_col + 1):
            cell = df.cell(row=row, column=column)
            if column == start_col and cell.value == name:
                check = True
                break
            if column == end_col and cell.value == link:
                check = True
                break
        if check == True:
            break
    return check


def check_if_release_date_is_wrong(value):
    check = False
    start_date = 2018
    end_date = datetime.datetime.now().year
    if len(value) == 0 or not value.isnumeric() or int(value) < start_date or int(value) > end_date:
        check = True
    return check


def check_the_value_of_all_text_boxes(file_path):
    check = True
    if check_empty_values():
        check = False
    if check_if_name_and_link_repeated(file_path, values[0].get(), values[1].get()):
        check = False
    return check


def edit_style_for_new_data(file_path, values):
    # Create a new workbook
    workbook = load_workbook(file_path)

    # Get the active sheet
    sheet = workbook.active

    # Specify the column names
    column_names = values.copy()

    # Specify the range of columns and row index
    start_column = 1
    end_column = 9
    row_index = len(sheet['A']) + 1

    # Set the height of row 2 to 30
    row_dimensions = sheet.row_dimensions[row_index]
    row_dimensions.height = 30

    # Create a border style
    border = Border(
        top=Side(border_style="thick", color="000000"),
        bottom=Side(border_style="thick", color="000000"),
        left=Side(border_style="thick", color="000000"),
        right=Side(border_style="thick", color="000000")
    )

    # Set the new values for cells in the specified range
    review_checkbox = YES if review_paper_check_value.get() == True else NO
    code_checkbox = YES if code_check_value.get() == True else NO
    for column, value in zip(range(start_column, end_column + 1), values):
        cell = sheet.cell(row=row_index, column=column)
        if isinstance(value, bool):
            value = "YES" if value == True else "NO"
        
        cell.value = value
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        if review_checkbox and code_checkbox:
            cell.fill = PatternFill(start_color=review_and_code_excel_cell_color, end_color=review_and_code_excel_cell_color, fill_type="solid")
        elif review_checkbox:
            cell.fill = PatternFill(start_color=review_only_excel_cell_color, end_color=review_only_excel_cell_color, fill_type="solid")
        elif code_checkbox:
            cell.fill = PatternFill(start_color=code_only_excel_cell_color, end_color=code_only_excel_cell_color, fill_type="solid")
        else:
            cell.fill = PatternFill(start_color=excel_cell_color, end_color=excel_cell_color, fill_type="solid")
        

    # Save the workbook to the specified file path
    workbook.save(file_path)


def check_research_paper_button():
    global is_paper_checked
    if not is_paper_checked:  # Paper hasn't been Checked yet
        if check_if_release_date_is_wrong(values[5].get()):
            check_empty_values()
            checkResearchPaper['foreground'] = error
            return
        file_path = create_saving_location()
        if not check_if_excel_file_exists(file_path):
            create_new_excel_file(file_path)
        if not check_the_value_of_all_text_boxes(file_path):
            checkResearchPaper['foreground'] = error
            print("Check Failed")
            return
        is_paper_checked = True
        checkResearchPaper['foreground'] = correct_check
    else:  # Paper has been Checked
        pass


def clear_all_fields():
    for value in values:
        value.delete(0, END)
        value.config(bg=textbox_background_color)
    review_paper_check_value.set(False)
    code_check_value.set(False)
    notesTextbox.insert(END, 'Nothing')
    checkResearchPaper.config(fg=normal_check_text)


def add_research_paper_button():
    global is_paper_checked
    if is_paper_checked:
        values = [paperNameTextbox.get().strip(), paperLinkTextbox.get().strip(),
                  modelAccuracyTextbox.get().strip(), modelDatasetTextbox.get().strip(),
                  modelAlgorithmTextbox.get().strip(), paperReleaseDateTextbox.get().strip(),
                  notesTextbox.get().strip(), review_paper_check_value.get(), code_check_value.get()]
        file_path = saving_file_path
        edit_style_for_new_data(file_path, values)
        print("ok")
        clear_all_fields()
        is_paper_checked = False


def task():
    loading_screen.destroy()
    main_window.geometry(main_window_size)
    review_paper_check.place(x=130, y=670)
    code_check.place(x=270, y=670)


def window_center(win):
    """
    centers a tkinter window
    :param win: the main window or Toplevel window to center
    """
    win.update_idletasks()
    width = win.winfo_width()
    frm_width = win.winfo_rootx() - win.winfo_x()
    win_width = width + 2 * frm_width
    height = win.winfo_height()
    titlebar_height = win.winfo_rooty() - win.winfo_y()
    win_height = height + titlebar_height + frm_width
    x = win.winfo_screenwidth() // 2 - win_width // 2
    y = win.winfo_screenheight() // 2 - win_height // 2
    win.geometry('{}x{}+{}+{}'.format(width, height, x, y))
    win.deiconify()


##################################################################################################################

# Classes
class ImageLabel(tk.Label):
    """
    A Label that displays images, and plays them if they are gifs
    :im: A PIL Image instance or a string filename
    """

    def load(self, im):
        if isinstance(im, str):
            im = Image.open(im)
        frames = []

        try:
            for i in count(1):
                frames.append(ImageTk.PhotoImage(im.copy().resize((int(main_window_size[:3]), int(main_window_size[4:])))))
                im.seek(i)
        except EOFError:
            pass
        self.frames = cycle(frames)

        try:
            self.delay = im.info['duration']
        except:
            self.delay = 100

        if len(frames) == 1:
            self.config(image=next(self.frames))
        else:
            self.next_frame()

    def unload(self):
        self.config(image=None)
        self.frames = None

    def next_frame(self):
        if self.frames:
            self.config(image=next(self.frames))
            self.after(self.delay, self.next_frame)


##################################################################################################################

# Main Window and Loading Screen:
main_window = tk.Tk()
icon_img = PhotoImage(file=icon_img_path)
main_window.tk.call('wm', 'iconphoto', main_window._w, icon_img)
main_window.resizable(0,0)
main_window.title("Research Paper Ordering")
main_window['background'] = main_window_background_color
main_window.geometry(main_window_size)
window_center(main_window)
loading_screen = ImageLabel(main_window)
loading_screen.pack()
loading_screen.load(loading_screen_image)
main_window.after(loading_time, task)

##################################################################################################################

# Title
title_font = font.Font(size=24)
title = Label(main_window, text="Research Paper Ordering", font=title_font)
title["background"] = title_background_color
title["foreground"] = title_foreground_color
title.pack(pady=20)

##################################################################################################################

# Labels and Text Boxes
# First
label_and_textbox_1 = {
    "label_size": "15",
    "label_text": "Research Paper Name",
    "label_background_color": label_background_color,
    "label_foreground_color": label_foreground_color,
    "textbox_background_color": textbox_background_color
}
paperNameLabel = create_label(label_and_textbox_1)
paperNameLabel.pack(pady=0)
paperNameTextbox = create_textbox(label_and_textbox_1)
text_1 = tk.StringVar(main_window)
paperNameTextbox["textvariable"] = text_1
text_1.trace('w', paperNameTextbox_background_color)
paperNameTextbox.pack(padx=10, pady=5)

# Second
label_and_textbox_2 = {
    "label_size": "15",
    "label_text": "Research Paper Link",
    "label_background_color": label_background_color,
    "label_foreground_color": label_foreground_color,
    "textbox_background_color": textbox_background_color
}
paperLinkLabel = create_label(label_and_textbox_2)
paperLinkLabel.pack(pady=5)
paperLinkTextbox = create_textbox(label_and_textbox_2)
text_2 = tk.StringVar(main_window)
paperLinkTextbox["textvariable"] = text_2
text_2.trace('w', paperLinkTextbox_background_color)
paperLinkTextbox.pack(padx=10, pady=5)

# Third
label_and_textbox_3 = {
    "label_size": "15",
    "label_text": "Model Accuracy",
    "label_background_color": label_background_color,
    "label_foreground_color": label_foreground_color,
    "textbox_background_color": textbox_background_color
}
modelAccuracyLabel = create_label(label_and_textbox_3)
modelAccuracyLabel.pack(pady=5)
modelAccuracyTextbox = create_textbox(label_and_textbox_3)
text_3 = tk.StringVar(main_window)
modelAccuracyTextbox["textvariable"] = text_3
text_3.trace('w', modelAccuracyTextbox_background_color)
modelAccuracyTextbox.pack(padx=10, pady=5)

# Fourth
label_and_textbox_4 = {
    "label_size": "15",
    "label_text": "Model Dataset",
    "label_background_color": label_background_color,
    "label_foreground_color": label_foreground_color,
    "textbox_background_color": textbox_background_color
}
modelDatasetLabel = create_label(label_and_textbox_4)
modelDatasetLabel.pack(pady=5)
modelDatasetTextbox = create_textbox(label_and_textbox_4)
text_4 = tk.StringVar(main_window)
modelDatasetTextbox["textvariable"] = text_4
text_4.trace('w', modelDatasetTextbox_background_color)
modelDatasetTextbox.pack(padx=10, pady=5)

# Fifth
label_and_textbox_5 = {
    "label_size": "15",
    "label_text": "Model Algorithm",
    "label_background_color": label_background_color,
    "label_foreground_color": label_foreground_color,
    "textbox_background_color": textbox_background_color
}
modelAlgorithmLabel = create_label(label_and_textbox_5)
modelAlgorithmLabel.pack(pady=5)
modelAlgorithmTextbox = create_textbox(label_and_textbox_5)
text_5 = tk.StringVar(main_window)
modelAlgorithmTextbox["textvariable"] = text_5
text_5.trace('w', modelAlgorithmTextbox_background_color)
modelAlgorithmTextbox.pack(padx=10, pady=5)

# Sixth
label_and_textbox_6 = {
    "label_size": "15",
    "label_text": "Paper Release Date",
    "label_background_color": label_background_color,
    "label_foreground_color": label_foreground_color,
    "textbox_background_color": textbox_background_color
}
paperReleaseDateLabel = create_label(label_and_textbox_6)
paperReleaseDateLabel.pack(pady=5)
paperReleaseDateTextbox = create_textbox(label_and_textbox_6)
text_6 = tk.StringVar(main_window)
paperReleaseDateTextbox["textvariable"] = text_6
text_6.trace('w', paperReleaseDateTextbox_background_color)
paperReleaseDateTextbox.pack(padx=10, pady=5)

# Seventh
label_and_textbox_7 = {
    "label_size": "15",
    "label_text": "Notes",
    "label_background_color": label_background_color,
    "label_foreground_color": label_foreground_color,
    "textbox_background_color": textbox_background_color
}
notesLabel = create_label(label_and_textbox_7)
notesLabel.pack(pady=5)
notesTextbox = create_textbox(label_and_textbox_7)
text_7 = tk.StringVar(main_window)
notesTextbox["textvariable"] = text_7
text_7.trace('w', notesTextbox_background_color)
notesTextbox.insert(END, 'Nothing')
notesTextbox.pack(padx=10, pady=5)

##################################################################################################################

# CheckBoxs

# Review Paper check
review_paper_check_value = tk.BooleanVar()
def onCheck():
	print(f"Is radiobutton checked? {review_paper_check_value.get()}")
review_paper_check = tk.Checkbutton(
	master=main_window,
    text="Review Paper",
    variable=review_paper_check_value,
    onvalue=True,
    offvalue=False,
    command=onCheck,
    background=paper_checkbox_background_color,
    foreground=paper_checkbox_foreground_color
)



# Code check
code_check_value = tk.BooleanVar()
def onCheck():
	print(f"Is radiobutton checked? {code_check_value.get()}")
code_check = tk.Checkbutton(
	master=main_window,
    text="Code",
    variable=code_check_value,
    onvalue=True,
    offvalue=False,
    command=onCheck,
    background=code_checkbox_background_color,
    foreground=code_checkbox_foreground_color
)


##################################################################################################################

# Buttons

button_font = font.Font(size=button_font_size)
checkResearchPaper = Button(main_window, text="Check Research Paper",
                            command=check_research_paper_button,
                            font=button_font)
checkResearchPaper['background'] = check_button_background_color
checkResearchPaper['foreground'] = check_button_foreground_color
checkResearchPaper.pack(padx=35, pady=10, side=LEFT)

addResearchPaper = Button(main_window, text="Add Research Paper", command=add_research_paper_button,
                          font=button_font)
addResearchPaper['background'] = add_button_background_color
addResearchPaper['foreground'] = add_button_foreground_color
addResearchPaper.pack(padx=30, pady=10, side=RIGHT)

##################################################################################################################

values = [paperNameTextbox, paperLinkTextbox, modelAccuracyTextbox,
          modelDatasetTextbox, modelAlgorithmTextbox, paperReleaseDateTextbox, notesTextbox]

##################################################################################################################

main_window.mainloop()
